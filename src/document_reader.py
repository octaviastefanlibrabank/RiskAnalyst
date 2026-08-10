"""
Reads the raw input documents (PDF, XLSM/XLSX, and legacy-named .doc that are
actually saved HTML) into a uniform RawDocument representation.

Nothing here modifies the original files - workbooks are opened read-only and
never saved back.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pymupdf  # PyMuPDF
from bs4 import BeautifulSoup

from .models import RawDocument

# Filename markers used to classify a company's input files.
# The KO/output files are intentionally NOT matched here as "input" -
# see classify_document() and discover_company_documents().
_REFERAT_MARKERS = ["referat"]
_CRC_MARKERS = ["scor crc", "crc", "verificari"]
_GROUP_FIN_MARKERS = ["grup"]
_FIN_MARKERS = ["analiza_financiara", "analiza financiara"]


def list_companies(data_root: Path) -> list[str]:
    """Return company folder names found under data_root."""
    if not data_root.exists():
        return []
    return sorted(p.name for p in data_root.iterdir() if p.is_dir())


def find_company_dir(data_root: Path, company: str) -> Path | None:
    """Case-insensitive, whitespace-tolerant match of a company name to its folder."""
    target = " ".join(company.split()).lower()
    for p in data_root.iterdir():
        if p.is_dir() and " ".join(p.name.split()).lower() == target:
            return p
    return None


def classify_document(filename: str) -> str:
    name = filename.lower()
    if name.startswith("output"):
        return "output_reference"  # NEVER used as input for the same company
    if not name.startswith("input"):
        return "other"
    if any(m in name for m in _CRC_MARKERS):
        return "crc_pdf"
    if any(m in name for m in _REFERAT_MARKERS):
        return "referat_pdf"
    if any(m in name for m in _GROUP_FIN_MARKERS) and name.endswith((".xlsm", ".xlsx")):
        return "financial_group_xlsm"
    if any(m in name for m in _FIN_MARKERS) and name.endswith((".xlsm", ".xlsx")):
        return "financial_xlsm"
    if name.endswith(".pdf"):
        return "pdf_other"
    if name.endswith((".xlsm", ".xlsx")):
        return "spreadsheet_other"
    return "other"


def sniff_doc_format(path: Path) -> str:
    """.doc files in this dataset are HTML saved with a .doc extension. Verify before parsing."""
    head = path.open("rb").read(512)
    text_head = head.decode("utf-8", errors="ignore").lstrip().lower()
    if text_head.startswith("<html") or "<!doctype" in text_head or "<meta" in text_head:
        return "html"
    if head.startswith(b"PK"):
        return "docx_zip"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "doc_binary"
    return "unknown"


def read_pdf(path: Path) -> RawDocument:
    doc = pymupdf.open(path)
    parts = []
    for i, page in enumerate(doc):
        parts.append(f"\n--- page {i + 1} ---\n" + page.get_text())
    text = "".join(parts)
    doc.close()
    return RawDocument(
        filename=path.name,
        doc_type=classify_document(path.name),
        text=text,
        char_count=len(text),
    )


def read_html_doc(path: Path) -> RawDocument:
    """Reads a saved-as-.doc HTML file (used for OUTPUT reference files during evaluation)."""
    fmt = sniff_doc_format(path)
    if fmt != "html":
        return RawDocument(
            filename=path.name,
            doc_type=classify_document(path.name),
            text="",
            warnings=[f"Unsupported .doc format detected ({fmt}); treated as reference-only, not parsed."],
        )
    html = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "lxml")
    text = "\n".join(line.strip() for line in soup.get_text(separator="\n").split("\n") if line.strip())
    return RawDocument(
        filename=path.name,
        doc_type=classify_document(path.name),
        text=text,
        char_count=len(text),
    )


def read_workbook(path: Path) -> RawDocument:
    """
    Reads all sheets of an xlsx/xlsm workbook. Uses cached computed values
    (data_only=True) where available, and falls back to the raw formula
    string (data_only=False) for cells whose cached value is missing
    (e.g. the workbook was never recalculated/opened in Excel).
    Read-only - the workbook is never modified or saved.
    """
    warnings: list[str] = []
    try:
        wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_vba=False)
    except Exception as e:
        warnings.append(f"Could not open with data_only=True ({e}); retrying without keep_vba.")
        wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)

    wb_formulas = None
    try:
        wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception as e:
        warnings.append(f"Could not open with data_only=False ({e}); formula fallback disabled.")

    sheets: dict[str, list[list]] = {}
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name] if wb_formulas and sheet_name in wb_formulas.sheetnames else None
        rows: list[list] = []
        # Read-only worksheets only support fast sequential access (iter_rows), not
        # random .cell(row, col) lookups (that degrades to O(n) per call -> O(n^2)
        # overall on wide sheets). Zip the two row iterators in lockstep instead.
        formula_rows_iter = ws_f.iter_rows() if ws_f is not None else iter(())
        for row, formula_row in zip(ws_v.iter_rows(), formula_rows_iter):
            row_vals = []
            for c_idx, cell in enumerate(row):
                val = cell.value
                if val is None and c_idx < len(formula_row):
                    fval = formula_row[c_idx].value
                    if isinstance(fval, str) and fval.startswith("="):
                        val = fval  # keep as formula marker; no cached value existed
                row_vals.append(val)
            if any(v is not None for v in row_vals):
                rows.append(row_vals)
        if rows:
            sheets[sheet_name] = rows

    wb_values.close()
    if wb_formulas:
        wb_formulas.close()

    return RawDocument(
        filename=path.name,
        doc_type=classify_document(path.name),
        sheets=sheets,
        char_count=sum(len(r) for rows in sheets.values() for r in rows),
        warnings=warnings,
    )


def read_any(path: Path) -> RawDocument:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_workbook(path)
    if suffix == ".doc":
        return read_html_doc(path)
    return RawDocument(
        filename=path.name,
        doc_type="other",
        warnings=[f"No reader implemented for extension {suffix!r}."],
    )


def discover_company_documents(company_dir: Path, include_output: bool = False) -> list[RawDocument]:
    """
    Reads every INPUT file for a company. OUTPUT reference files are excluded
    by default (they must never be used as input for the same company - see
    project spec section 2 on data leakage). Pass include_output=True only for
    evaluation purposes (evaluate.py).
    """
    docs = []
    for path in sorted(company_dir.iterdir()):
        if not path.is_file():
            continue
        doc_type = classify_document(path.name)
        if doc_type == "output_reference" and not include_output:
            continue
        docs.append(read_any(path))
    return docs
